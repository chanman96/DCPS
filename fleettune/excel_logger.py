"""
Telemetry Excel logger.

Appends one row per vehicle every LOG_INTERVAL_S seconds of *simulated* time (so the log
cadence tracks perceived vehicle time, not wall clock, even when --time-scale speeds the
sim up), with a full ECU + driver snapshot and a UTC timestamp. One worksheet per vehicle.

The workbook is kept in memory and flushed to disk periodically (not on every row — xlsx
writes aren't cheap) plus on stop() and on-demand via flush().
"""
from __future__ import annotations
import threading
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

LOG_INTERVAL_S = 2.0

HEADER = [
    "timestamp_utc", "vehicle_id", "driver_name", "driver_profile",
    "lat", "lng", "heading_deg",
    "engine_rpm", "current_gear", "vehicle_speed_kph", "engine_load_pct", "accel_pedal_pct",
    "coolant_temp_c", "oil_temp_c", "fuel_temp_c", "intake_manifold_temp_c",
    "exhaust_gas_temp_c", "boost_pressure_kpa", "air_inlet_pressure_kpa",
    "fuel_delivery_pressure_kpa", "engine_oil_pressure_kpa", "engine_oil_level_pct",
    "fuel_rate_lph", "fuel_economy_kmpl", "inst_fuel_economy_kmpl",
    "fuel_level_pct", "total_vehicle_distance_km", "trip_distance_km",
    "battery_voltage", "ignition_on", "active_dtc_count",
    "fatigue_score", "continuous_drive_h", "driver_score",
    "harsh_brake_count", "harsh_accel_count", "idle_minutes", "idle_fuel_l",
]


class ExcelLogger:
    """Buffers rows in memory and periodically flushes to an .xlsx file — one sheet per vehicle."""

    def __init__(self, out_dir: str | Path = "logs"):
        self.out_dir = Path(out_dir)
        self.path: Path | None = None
        self.wb: Workbook | None = None
        self.sheets: dict[str, object] = {}
        self.row_count = 0
        self.started_at: float | None = None
        self._running = False
        self._lock = threading.Lock()

    def start(self) -> Path:
        self.out_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        with self._lock:
            self.path = self.out_dir / f"fleet_telemetry_{ts}.xlsx"
            self.wb = Workbook()
            self.wb.remove(self.wb.active)   # sheets are created lazily, per vehicle
            self.sheets = {}
            self.row_count = 0
            self.started_at = time.time()
            self._running = True
        return self.path

    def stop(self):
        self._running = False
        self.flush()

    def is_running(self) -> bool:
        return self._running

    def log_snapshot(self, vid: str, row: list):
        with self._lock:
            if not self._running or self.wb is None:
                return
            ws = self.sheets.get(vid)
            if ws is None:
                ws = self.wb.create_sheet(title=vid[:31])   # xlsx sheet-name length cap
                ws.append(HEADER)
                self.sheets[vid] = ws
            ws.append(row)
            self.row_count += 1

    def flush(self) -> Path | None:
        with self._lock:
            if self.wb is None or self.path is None:
                return None
            if not self.wb.sheetnames:
                self.wb.create_sheet(title="Sheet1")
            self.wb.save(self.path)
            return self.path

    def status(self) -> dict:
        return {
            "running": self._running,
            "path": str(self.path) if self.path else None,
            "rows": self.row_count,
            "vehicles": len(self.sheets),
            "started_at": self.started_at,
            "interval_s": LOG_INTERVAL_S,
        }
